import functions_framework
from google.cloud import secretmanager
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import numpy as np
import re

import google.auth.transport.requests
import google.oauth2.id_token
import requests

import io
import os
import json
import base64
from datetime import datetime, timedelta
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd


def get_secret(project, secret_name):
    """
    Retrieve a secret from Google Cloud Secret Manager.

    Args:
        project (str): The GCP project ID.
        secret_name (str): The name of the secret.

    Returns:
        str: The secret value.
    """
    client = secretmanager.SecretManagerServiceClient()
    name = f"projects/{project}/secrets/{secret_name}/versions/latest"
    response = client.access_secret_version(name=name)
    return response.payload.data.decode("UTF-8")

secret_name = "extractor-key"
project = "innovacion-402319"

credentials_json = get_secret(project, secret_name)
credentials = service_account.Credentials.from_service_account_info(
    json.loads(credentials_json), scopes=["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/spreadsheets"]
)
service = build("drive", "v3", credentials=credentials)
sheets_service = build('sheets', 'v4', credentials=credentials)

def create_local_folders():
    """
    Create the necessary local folders for processing.
    """
    create_tmp_folder("trazabilidad")
    create_tmp_folder("resultados")
    create_tmp_folder("libro_no_conciliados")
    create_tmp_folder("banco_no_conciliados")
    create_tmp_folder("libro_bruto")
    create_tmp_folder("banco_bruto")
    create_tmp_folder("alerta")
    

def create_tmp_folder(folder_name):
    """
    Create a temporary folder if it does not exist.

    Args:
        folder_name (str): The name of the folder.
    """
    newpath = f'./tmp/{folder_name}' 
    if not os.path.exists(newpath):
        os.makedirs(newpath)

def get_input_data(request):
    """
    Extract input data from an HTTP request.

    Args:
        request: The HTTP request object.

    Returns:
        tuple: A tuple containing file_id1, file_id2, month, year, book_id.
    """
    request_json = request.get_json(silent=True)
    archivo_libro = request_json["archivo_libro"]
    archivo_banco = request_json["archivo_banco"]
    month = request_json["month"]
    year = request_json["year"]
    book_id = request_json["id"]
    return archivo_libro, archivo_banco, month, year, book_id

def download_file_from_appsheet(url, output_path):
    """
    Download a file from AppSheet.

    Args:
        url (str): The URL to download the file from.
        output_path (str): The path to save the downloaded file.
    """
    response = requests.get(url)
    if response.status_code == 200:
        with open(output_path, 'wb') as file:
            file.write(response.content)
        print(f"File downloaded successfully to {output_path}")
    else:
        print(f"Failed to download file. Status code: {response.status_code}")

def cargar_y_limpiar_datos(archivo_libro, archivo_banco):
    df_banco = pd.read_excel(archivo_banco)
    df_libro = pd.read_excel(archivo_libro, header=None)
    def es_debito(tipo):
        return bool(re.match(r'^d[ée]bit[oe]?$', tipo.lower()))

    header_row = None
    for i, row in df_libro.iterrows():
        if 'Cuenta Bancaria' in str(row.values):
            header_row = i
            break

    if header_row is not None:
        df_libro.columns = df_libro.iloc[header_row]
        df_libro = df_libro.iloc[header_row + 1:].reset_index(drop=True)
    else:
        raise ValueError("No se pudo encontrar la fila de encabezados en el archivo del libro")

    df_libro = df_libro.dropna(how='all')
    df_banco = df_banco.dropna(how='all')

    df_libro = df_libro.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    df_banco = df_banco.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    if 'Monto' in df_libro.columns:
        df_libro['Monto'] = pd.to_numeric(df_libro['Monto'], errors='coerce')
    if 'Monto' in df_banco.columns:
        df_banco['Monto'] = pd.to_numeric(df_banco['Monto'], errors='coerce')

    if 'Cuenta Bancaria' in df_libro.columns:
        df_libro['Cuenta Bancaria'] = df_libro['Cuenta Bancaria'].astype(str).str.replace(r'\D', '', regex=True)
    if 'Cuenta Bancaria' in df_banco.columns:
        df_banco['Cuenta Bancaria'] = df_banco['Cuenta Bancaria'].astype(str).str.replace(r'\D', '', regex=True)
    
    # Ajustar los montos según el tipo de transacción
    if 'Tipo' in df_libro.columns:
        df_libro['Monto'] = df_libro.apply(lambda row: -row['Monto'] if es_debito(str(row['Tipo'])) else row['Monto'], axis=1)
    if 'Tipo' in df_banco.columns:
        df_banco['Monto'] = df_banco.apply(lambda row: -row['Monto'] if es_debito(str(row['Tipo'])) else row['Monto'], axis=1)


    return df_libro, df_banco

def crear_referencia_2(referencia):
    if isinstance(referencia, str) and len(referencia) >= 7:
        if referencia[1] in ("C", "D", "E"):
            return referencia[2:5] + referencia[-4:-1]
        else:
            return referencia[1:4] + referencia[-3:]
    return ''


def crear_tipo_de_tarjeta(referencia):
    if isinstance(referencia, str) and len(referencia) >= 7:
        if referencia[1] in ("C", "D", "E"):
            return referencia[1]
        else:
            return referencia[0]
    return ''

def crear_tienda(referencia):
    if isinstance(referencia, str) and len(referencia) >= 7:
        if referencia[1] in ("C", "D", "E"):
            return referencia[2:5]
        else:
            return referencia[1:4]
    return ''

def crear_lote(referencia):
    if isinstance(referencia, str) and len(referencia) >= 7:
        if referencia[1] in ("C", "D", "E"):
            return referencia[-4:-1]
        else:
            return referencia[-3:]
    return ''

def ajustar_monto(row, origen):
    monto = row['Monto']
    cuenta = str(row['Cuenta Bancaria'] if origen == 'libro' else row['Cuenta Bancaria'])
    referencia = str(row['Numero de Transacción'] if origen == 'libro' else row['Referencia'])

    comision = 0
    impuesto = 0
    porcentaje_comision = "0%"
    porcentaje_impuesto = "0%"

    if '1682' in cuenta:
        if origen == 'libro':
            comision = np.round(monto, 2) * 0.001
            monto *= 1.001  # Añadir 0.10%
            porcentaje_comision = "0.10%"

        if any(referencia[i] in ["C", "c"] for i in range(3)):
            impuesto = np.round(monto, 2) * 0.0431
            monto *= 1.0431  # Añadir 4.31%
            porcentaje_impuesto = "4.31%"

    # Guardar los valores calculados en el DataFrame
    row['Monto_Ajustado'] = np.round(monto, 2)
    row['Comision'] = comision
    row['Impuesto'] = impuesto
    row['Porcentaje_Comision'] = porcentaje_comision
    row['Porcentaje_Impuesto'] = porcentaje_impuesto

    return row

def marcar_conciliados(df, indices, columna_referencia):
    df.loc[indices, 'Conciliado'] = True
    return df

def conciliar_por_referencia(df_libro, df_banco, resultados, trazabilidad, tolerancia=0.05):
    conteo_libro = 0
    conteo_banco = 0
    for idx_libro, partida_libro in df_libro[~df_libro['Conciliado']].iterrows():
        # Primero intentamos conciliar sin tolerancia
        partidas_banco = df_banco[
            (df_banco['Referencia_2'] == partida_libro['Referencia_2']) &
            (df_banco['Monto_Ajustado'] == partida_libro['Monto_Ajustado']) &
            (df_banco['Fecha_Efectiva'].between(partida_libro['Fecha_Contable'] - timedelta(days=5), partida_libro['Fecha_Contable'] + timedelta(days=5))) &  # Comparar fechas dentro de ±5 días
            (~df_banco['Conciliado'])
        ]

        if partidas_banco.empty:
            # Si no hay coincidencias exactas, intentamos con tolerancia
            partidas_banco = df_banco[
                (df_banco['Referencia_2'] == partida_libro['Referencia_2']) &
                (abs(df_banco['Monto_Ajustado'] + partida_libro['Monto_Ajustado']) <= tolerancia) &
                (df_banco['Fecha_Efectiva'].between(partida_libro['Fecha_Contable'] - timedelta(days=5), partida_libro['Fecha_Contable'] + timedelta(days=5))) &  # Comparar fechas dentro de ±5 días
                (~df_banco['Conciliado'])
            ]

        if not partidas_banco.empty:
            partida_banco = partidas_banco.iloc[0]  # Tomar la primera coincidencia

            # Conciliación directa
            nueva_fila = pd.DataFrame({
                'Origen': ['Libro'],
                "Cuenta Bancaria": [partida_libro['Cuenta Bancaria']],
                "Cuenta": [str(110104)],
                "Subcuenta": ["0031" if int(partida_libro['Cuenta Bancaria']) % 10 == 2 else "0040"],
                "Descripcion": [partida_libro["Proveedor"]],
                "Tipo de tarjeta": [partida_libro["Tipo_de_tarjeta"]],
                "Tienda": [partida_libro["Tienda"]],
                "Lote": [partida_libro["Lote"]],
                "Referencia 2": [partida_libro["Referencia_2"]],
                'Referencia': [partida_libro['Numero de Transacción']],
                "Tipo": ["Debito"],
                "Monto": [partida_libro["Monto"]],
                "Monto2": [-abs(int(partida_libro["Monto"]))],
                "%comision": [partida_libro["Porcentaje_Comision"]],
                "comision": [partida_libro["Comision"]],
                "%impuesto": [partida_libro["Porcentaje_Impuesto"]],
                "impuesto": [partida_libro["Impuesto"]],
                'Monto_Ajustado': [partida_libro['Monto_Ajustado']],
                'Fecha': [partida_libro['Fecha_Contable']]
            })
            resultados = pd.concat([resultados, nueva_fila], ignore_index=True)

            nueva_fila = pd.DataFrame({
                'Origen': ['Banco'],
                "Cuenta Bancaria": [partida_banco['Cuenta Bancaria']],
                "Cuenta": [partida_banco['Cuenta Contable']],
                "Subcuenta": [partida_banco['Sub Cuenta']],
                "Descripcion": [partida_banco["Descripción"]],
                "Tipo de tarjeta": [partida_banco["Tipo_de_tarjeta"]],
                "Tienda": [partida_banco["Tienda"]],
                "Lote": [partida_banco["Lote"]],
                "Referencia 2": [partida_banco["Referencia_2"]],
                'Referencia': [partida_banco['Referencia']],
                "Tipo": ["Credito"],
                "Monto": [partida_banco["Monto"]],
                "Monto2": [abs(int(partida_banco["Monto"]))],
                "%comision": [partida_banco["Porcentaje_Comision"]],
                "comision": [partida_banco["Comision"]],
                "%impuesto": [partida_banco["Porcentaje_Impuesto"]],
                "impuesto": [partida_banco["Impuesto"]],
                'Monto_Ajustado': [partida_banco['Monto_Ajustado']],
                'Fecha': [partida_banco['Fecha_Efectiva']],
                "Banco": [partida_banco['Banco']]
            })
            resultados = pd.concat([resultados, nueva_fila], ignore_index=True)

            tipo_conciliacion = 'Referencia_2_Directa' if df_banco['Monto_Ajustado'].equals(partida_libro['Monto_Ajustado']) else 'Referencia_2_Directa_Tolerancia'
            nueva_fila2 = pd.DataFrame({
                'Partida_Libro': [partida_libro['Numero de Transacción']],
                'Partida_Banco': [partida_banco['Referencia']],
                'Tipo_Conciliacion': [tipo_conciliacion]
            })
            trazabilidad = pd.concat([trazabilidad, nueva_fila2], ignore_index=True)
            conteo_libro += 1
            conteo_banco += 1

            df_libro = marcar_conciliados(df_libro, [idx_libro], 'Numero de Transacción')
            df_banco = marcar_conciliados(df_banco, [partida_banco.name], 'Referencia')

    return resultados, df_libro, df_banco, trazabilidad, conteo_libro, conteo_banco

def conciliar_multiple_banco(df_libro, df_banco, resultados, trazabilidad):
    conteo_libro = 0
    conteo_banco = 0
    for idx_libro, partida_libro in df_libro[~df_libro['Conciliado']].iterrows():
        partidas_banco = df_banco[
            (df_banco['Referencia_2'] == partida_libro['Referencia_2']) &
            (~df_banco['Conciliado'])
        ]

        if len(partidas_banco) > 1:
            monto_libro = partida_libro['Monto_Ajustado']

            # Probar todas las combinaciones posibles de partidas del banco
            for r in range(2, len(partidas_banco) + 1):
                for combo in combinations(partidas_banco.index, r):
                    suma_montos_banco = df_banco.loc[list(combo), 'Monto_Ajustado'].sum()

                    if abs(suma_montos_banco + monto_libro) <= 0.05:
                        nueva_fila_libro = pd.DataFrame({
                            'Origen': ['Libro'],
                            "Cuenta Bancaria": [partida_libro['Cuenta Bancaria']],
                            "Cuenta": [str(110104)],
                            "Subcuenta": ["0031" if int(partida_libro['Cuenta Bancaria']) % 10 == 2 else "0040"],
                            "Descripcion": [partida_libro["Proveedor"]],
                            "Tipo de tarjeta": [partida_libro["Tipo_de_tarjeta"]],
                            "Tienda": [partida_libro["Tienda"]],
                            "Lote": [partida_libro["Lote"]],
                            "Referencia 2": [partida_libro["Referencia_2"]],
                            'Referencia': [partida_libro['Numero de Transacción']],
                            "Tipo": ["Debito"],
                            "Monto": [partida_libro["Monto"]],
                            "Monto2": [-abs(int(partida_libro["Monto"]))],
                            "%comision": [partida_libro["Porcentaje_Comision"]],
                            "comision": [partida_libro["Comision"]],
                            "%impuesto": [partida_libro["Porcentaje_Impuesto"]],
                            "impuesto": [partida_libro["Impuesto"]],
                            'Monto_Ajustado': [partida_libro['Monto_Ajustado']],
                            'Fecha': [partida_libro['Fecha_Contable']]
                        })
                        resultados = pd.concat([resultados, nueva_fila_libro], ignore_index=True)

                        for idx_banco in combo:
                            partida_banco = df_banco.loc[idx_banco]
                            nueva_fila_banco = pd.DataFrame({
                                'Origen': ['Banco'],
                                "Cuenta Bancaria": [partida_banco['Cuenta Bancaria']],
                                "Cuenta": [partida_banco['Cuenta Contable']],
                                "Subcuenta": [partida_banco['Sub Cuenta']],
                                "Descripcion": [partida_banco["Descripción"]],
                                "Tipo de tarjeta": [partida_banco["Tipo_de_tarjeta"]],
                                "Tienda": [partida_banco["Tienda"]],
                                "Lote": [partida_banco["Lote"]],
                                "Referencia 2": [partida_banco["Referencia_2"]],
                                'Referencia': [partida_banco['Referencia']],
                                "Tipo": ["Credito"],
                                "Monto": [partida_banco["Monto"]],
                                "Monto2": [abs(int(partida_banco["Monto"]))],
                                "%comision": [partida_banco["Porcentaje_Comision"]],
                                "comision": [partida_banco["Comision"]],
                                "%impuesto": [partida_banco["Porcentaje_Impuesto"]],
                                "impuesto": [partida_banco["Impuesto"]],
                                'Monto_Ajustado': [partida_banco['Monto_Ajustado']],
                                'Fecha': [partida_banco['Fecha_Efectiva']],
                                "Banco": [partida_banco['Banco']]
                            })
                            resultados = pd.concat([resultados, nueva_fila_banco], ignore_index=True)

                        nueva_fila2 = pd.DataFrame({
                            'Partida_Libro': [partida_libro['Numero de Transacción']],
                            'Partida_Banco': [','.join(df_banco.loc[list(combo), 'Referencia'].astype(str))],
                            'Tipo_Conciliacion': ['Multiple Banco']
                        })
                        trazabilidad = pd.concat([trazabilidad, nueva_fila2], ignore_index=True)
                        conteo_libro += 1
                        conteo_banco += len(combo)
                        df_libro.at[idx_libro, 'Conciliado'] = True
                        df_banco.loc[list(combo), 'Conciliado'] = True
                        break  # Salir del bucle interno si se encuentra una coincidencia

                if df_libro.loc[idx_libro, 'Conciliado']:
                    break  # Salir del bucle externo si ya se concilió esta partida del libro

    return resultados, df_libro, df_banco, trazabilidad, conteo_banco, conteo_libro

def conciliar_multiple_libro(df_libro, df_banco, resultados, trazabilidad):
    conteo_libro = 0
    conteo_banco = 0
    for idx_banco, partida_banco in df_banco[~df_banco['Conciliado']].iterrows():
        partidas_libro = df_libro[
            (df_libro['Referencia_2'] == partida_banco['Referencia_2']) &
            (~df_libro['Conciliado'])
        ]
        if len(partidas_libro) == 2 and abs(partidas_libro['Monto_Ajustado'].sum() + partida_banco['Monto_Ajustado']) <= 0.05:
            for idx_libro in partidas_libro.index:
                nueva_fila_libro = pd.DataFrame({
                    'Origen': ['Libro'],
                    "Cuenta Bancaria": [df_libro.loc[idx_libro, 'Cuenta Bancaria']],
                    "Cuenta": [str(110104)],  # Agregado
                    "Subcuenta": ["0031" if int(df_libro.loc[idx_libro, 'Cuenta Bancaria']) % 10 == 2 else "0040"],
                    "Descripcion": [df_libro.loc[idx_libro, "Proveedor"]],
                    "Tipo de tarjeta": [df_libro.loc[idx_libro, "Tipo_de_tarjeta"]],
                    "Tienda": [df_libro.loc[idx_libro, "Tienda"]],
                    "Lote": [df_libro.loc[idx_libro, "Lote"]],
                    "Referencia 2": [df_libro.loc[idx_libro, "Referencia_2"]],
                    'Referencia': [df_libro.loc[idx_libro, 'Numero de Transacción']],
                    "Tipo": ["Debito"],
                    "Monto": [df_libro.loc[idx_libro, "Monto"]],
                    "Monto2": [-abs(int(df_libro.loc[idx_libro, "Monto"]))],
                    "%comision": [df_libro.loc[idx_libro, "Porcentaje_Comision"]],
                    "comision": [df_libro.loc[idx_libro, "Comision"]],
                    "%impuesto": [df_libro.loc[idx_libro, "Porcentaje_Impuesto"]],
                    "impuesto": [df_libro.loc[idx_libro, "Impuesto"]],
                    'Monto_Ajustado': [df_libro.loc[idx_libro, 'Monto_Ajustado']],
                    'Fecha': [df_libro.loc[idx_libro, 'Fecha_Contable']],
                })
                resultados = pd.concat([resultados, nueva_fila_libro], ignore_index=True)

            nueva_fila_banco = pd.DataFrame({
                'Origen': ['Banco'],
                "Cuenta Bancaria": [partida_banco['Cuenta Bancaria']],
                "Cuenta": [partida_banco['Cuenta Contable']],
                "Subcuenta": [partida_banco['Sub Cuenta']],
                "Descripcion": [partida_banco["Descripción"]],
                "Tipo de tarjeta": [partida_banco["Tipo_de_tarjeta"]],
                "Tienda": [partida_banco["Tienda"]],
                "Lote": [partida_banco["Lote"]],
                "Referencia 2": [partida_banco["Referencia_2"]],
                'Referencia': [partida_banco['Referencia']],
                "Tipo": ["Credito"],
                "Monto": [partida_banco["Monto"]],
                "Monto2": [abs(int(partida_banco["Monto"]))],
                "%comision": [partida_banco["Porcentaje_Comision"]],
                "comision": [partida_banco["Comision"]],
                "%impuesto": [partida_banco["Porcentaje_Impuesto"]],
                "impuesto": [partida_banco["Impuesto"]],
                'Monto_Ajustado': [partida_banco['Monto_Ajustado']],
                'Fecha': [partida_banco['Fecha_Efectiva']],
                "Banco": [partida_banco['Banco']] 
            })
            resultados = pd.concat([resultados, nueva_fila_banco], ignore_index=True)

            nueva_fila2 = pd.DataFrame({
                'Partida_Libro': [','.join(map(str, partidas_libro))],
                'Partida_Banco': [partida_banco['Referencia']],
                'Tipo_Conciliacion': ['Multiple_Libro']
            })
            trazabilidad = pd.concat([trazabilidad, nueva_fila2], ignore_index=True)
            conteo_libro += len(partidas_libro)
            conteo_banco += 1

            df_libro = marcar_conciliados(df_libro, partidas_libro.index, 'Numero de Transacción')
            df_banco = marcar_conciliados(df_banco, [idx_banco], 'Referencia')
    print("+--+-+-+-+-",resultados)
    return resultados, df_libro, df_banco, trazabilidad, conteo_banco, conteo_libro

def conciliar_pagos(archivo_libro, archivo_banco):

    conteo_por_criterio = {
        'Criterio 1': {'libro': 0, 'banco': 0},
        'Criterio 2': {'libro': 0, 'banco': 0},
        'Criterio 3': {'libro': 0, 'banco': 0},
        'Criterio 4': {'libro': 0, 'banco': 0},
        'Criterio 5': {'libro': 0, 'banco': 0},
    }

    # Cargar y limpiar los datos
    df_libro, df_banco = cargar_y_limpiar_datos(archivo_libro, archivo_banco)

    # Añadir columna 'Conciliado'
    df_libro['Conciliado'] = False
    df_banco['Conciliado'] = False

    # Inicializar contadores
    conteo_libro_conciliado = 0
    conteo_banco_conciliado = 0

    total_banco_bruto = len(df_banco)

    # Filtrar registros del libro
    df_banco = df_banco[df_banco['Descripción'].str.contains('AB.LOTE|LIQUIDACI', na=False)]

    # Crear Referencia 2 y ajustar montos
    df_libro['Referencia_2'] = df_libro['Numero de Transacción'].apply(crear_referencia_2)
    df_banco['Referencia_2'] = df_banco['Referencia'].apply(crear_referencia_2)

    # Crear Referencia 2 y ajustar montos
    df_libro['Tipo_de_tarjeta'] = df_libro['Numero de Transacción'].apply(crear_tipo_de_tarjeta)
    df_banco['Tipo_de_tarjeta'] = df_banco['Referencia'].apply(crear_tipo_de_tarjeta)

    # Crear Referencia 2 y ajustar montos
    df_libro['Tienda'] = df_libro['Numero de Transacción'].apply(crear_tienda)
    df_banco['Tienda'] = df_banco['Referencia'].apply(crear_tienda)

    # Crear Referencia 2 y ajustar montos
    df_libro['Lote'] = df_libro['Numero de Transacción'].apply(crear_lote)
    df_banco['Lote'] = df_banco['Referencia'].apply(crear_lote)
    
    # Modificar cómo se aplica la función ajustar_monto
    df_libro = df_libro.apply(lambda row: ajustar_monto(row, 'libro'), axis=1)
    df_banco = df_banco.apply(lambda row: ajustar_monto(row, 'banco'), axis=1)

    df_libro['Fecha_Contable'] = pd.to_datetime(df_libro['Fecha Contable'])
    df_banco['Fecha_Efectiva'] = pd.to_datetime(df_banco['Fecha Efectiva'])

    resultados = pd.DataFrame(columns=[
     "Origen", "Cuenta Bancaria", "Cuenta", "Subcuenta", "Descripcion", 
    "Fecha", "Referencia", "Tipo de tarjeta", "Tienda", "Lote", 
    "Referencia 2", "Tipo", "Monto", "Monto2", 
    "%comision", "%impuesto", "comision", "impuesto", 
    "Monto_Ajustado"
    ])
    trazabilidad = pd.DataFrame(columns=['Partida_Libro', 'Partida_Banco', 'Tipo_Conciliacion'])
    alerta = pd.DataFrame(columns=['Partida_Libro', 'Partida_Banco', 'Tipo_Conciliacion'])

    # # Modificar las funciones de conciliación para que también actualicen la trazabilidad
    resultados, df_libro, df_banco, trazabilidad, conteo_libro, conteo_banco = conciliar_por_referencia(df_libro, df_banco, resultados, trazabilidad)

    resultados, df_libro, df_banco, trazabilidad, conteo_libro, conteo_banco = conciliar_multiple_banco(df_libro, df_banco, resultados, trazabilidad)
    conteo_por_criterio['Criterio 3']['libro'] = conteo_libro
    conteo_por_criterio['Criterio 3']['banco'] = conteo_banco

    resultados, df_libro, df_banco, trazabilidad, conteo_libro, conteo_banco = conciliar_multiple_libro(df_libro, df_banco, resultados, trazabilidad)
    conteo_por_criterio['Criterio 4']['libro'] = conteo_libro
    conteo_por_criterio['Criterio 4']['banco'] = conteo_banco

    resultados, df_libro, df_banco, trazabilidad, conteo_libro, conteo_banco = conciliar_por_criterios_similares(df_libro, df_banco, resultados, trazabilidad)
    conteo_por_criterio['Criterio 5']['libro'] = conteo_libro
    conteo_por_criterio['Criterio 5']['banco'] = conteo_banco

    df_libro, df_banco, alerta = buscar_alertas(df_libro, df_banco, alerta)

        # Contar cuántos registros hay de cada banco
    conteo_bancos = resultados['Banco'].value_counts()
    
    # Obtener el conteo específico de Banco A y Banco B
    conteo_banca_amiga = conteo_bancos.get('BANCAMIGA POS EXTERNO', 0)  # Devuelve 0 si no existe
    conteo_banco_fondo_comun = conteo_bancos.get('BANCO FONDO COMUN', 0)  # Devuelve 0 si no existe


    conteo_libro_conciliado = df_libro['Conciliado'].sum()
    conteo_banco_conciliado = df_banco['Conciliado'].sum()

    no_conciliados_libro = df_libro[~df_libro['Conciliado']]
    no_conciliados_banco = df_banco[~df_banco['Conciliado']]

    conteo_por_criterio['Criterio 1']['libro'] = conteo_libro_conciliado - (conteo_por_criterio['Criterio 5']['libro'] + conteo_por_criterio['Criterio 4']['libro'] + conteo_por_criterio['Criterio 3']['libro'] + conteo_por_criterio['Criterio 2']['libro'])
    conteo_por_criterio['Criterio 1']['banco'] = conteo_banco_conciliado - (conteo_por_criterio['Criterio 5']['banco'] + conteo_por_criterio['Criterio 4']['banco'] + conteo_por_criterio['Criterio 3']['banco'] + conteo_por_criterio['Criterio 2']['banco'])
    

    return resultados, conteo_libro_conciliado, conteo_banco_conciliado, df_libro, df_banco, no_conciliados_libro, no_conciliados_banco, trazabilidad, conteo_por_criterio, total_banco_bruto, alerta, conteo_banca_amiga, conteo_banco_fondo_comun

def referencias_similares(ref1, ref2, max_diferencias=2):

    # Extraer los primeros 3 dígitos (tienda) y los últimos 3 dígitos (lote)
    tienda1, lote1 = ref1[:3], ref1[-3:]
    tienda2, lote2 = ref2[:3], ref2[-3:]

    # Comparar solo los últimos 3 dígitos (lote)
    diferencias = sum(1 for a, b in zip(lote1, lote2) if a != b)

    # Si ambas referencias tienen tienda definida (no asteriscos), deben coincidir
    if diferencias <= max_diferencias:

        if tienda1 == tienda2:
            return 1
        else: 
            return 0
    else:
        return -1


def conciliar_por_criterios_similares(df_libro, df_banco, resultados, trazabilidad):
    conteo_libro = 0
    conteo_banco = 0
    
    for idx_libro, partida_libro in df_libro[~df_libro['Conciliado']].iterrows():
        partidas_banco = df_banco[
            (abs(df_banco['Monto_Ajustado'] + partida_libro['Monto_Ajustado']) <= 0.05) &
            (df_banco['Fecha_Efectiva'].between(partida_libro['Fecha_Contable'] - timedelta(days=5), partida_libro['Fecha_Contable'] + timedelta(days=5))) &
            (~df_banco['Conciliado'])
        ]

        for idx_banco, partida_banco in partidas_banco.iterrows():
            if referencias_similares(partida_libro['Referencia_2'], partida_banco['Referencia_2']) == 1:
                # Crear nueva fila para el libro
                nueva_fila_libro = pd.DataFrame({
                    'Origen': ['Libro'],
                    "Cuenta Bancaria": [partida_libro['Cuenta Bancaria']],
                    "Cuenta": [str(110104)],
                    "Subcuenta": ["0031" if int(partida_libro['Cuenta Bancaria']) % 10 == 2 else "0040"],
                    "Descripcion": [partida_libro["Proveedor"]],
                    "Tipo de tarjeta": [partida_libro["Tipo_de_tarjeta"]],
                    "Tienda": [partida_libro["Tienda"]],
                    "Lote": [partida_libro["Lote"]],
                    "Referencia 2": [partida_libro["Referencia_2"]],
                    'Referencia': [partida_libro['Numero de Transacción']],
                    "Tipo": ["Debito"],
                    "Monto": [partida_libro["Monto"]],
                    "Monto2": [-abs(int(partida_libro["Monto"]))],
                    "%comision": [partida_libro["Porcentaje_Comision"]],
                    "comision": [partida_libro["Comision"]],
                    "%impuesto": [partida_libro["Porcentaje_Impuesto"]],
                    "impuesto": [partida_libro["Impuesto"]],
                    'Monto_Ajustado': [partida_libro['Monto_Ajustado']],
                    'Fecha': [partida_libro['Fecha_Contable']]
                })
                resultados = pd.concat([resultados, nueva_fila_libro], ignore_index=True)

                # Crear nueva fila para el banco
                nueva_fila_banco = pd.DataFrame({
                    'Origen': ['Banco'],
                    "Cuenta Bancaria": [partida_banco['Cuenta Bancaria']],
                    "Cuenta": [partida_banco['Cuenta Contable']],
                    "Subcuenta": [partida_banco['Sub Cuenta']],
                    "Descripcion": [partida_banco["Descripción"]],
                    "Tipo de tarjeta": [partida_banco["Tipo_de_tarjeta"]],
                    "Tienda": [partida_banco["Tienda"]],
                    "Lote": [partida_banco["Lote"]],
                    "Referencia 2": [partida_banco["Referencia_2"]],
                    'Referencia': [partida_banco['Referencia']],
                    "Tipo": ["Credito"],
                    "Monto": [partida_banco["Monto"]],
                    "Monto2": [abs(int(partida_banco["Monto"]))],
                    "%comision": [partida_banco["Porcentaje_Comision"]],
                    "comision": [partida_banco["Comision"]],
                    "%impuesto": [partida_banco["Porcentaje_Impuesto"]],
                    "impuesto": [partida_banco["Impuesto"]],
                    'Monto_Ajustado': [partida_banco['Monto_Ajustado']],
                    'Fecha': [partida_banco['Fecha_Efectiva']],
                    "Banco": [partida_banco['Banco']]
                })
                resultados = pd.concat([resultados, nueva_fila_banco], ignore_index=True)

                # Agregar a la trazabilidad
                nueva_fila_trazabilidad = pd.DataFrame({
                    'Partida_Libro': [partida_libro['Numero de Transacción']],
                    'Partida_Banco': [partida_banco['Referencia']],
                    'Tipo_Conciliacion': ['conciliar_por_referencias_similares']
                })
                trazabilidad = pd.concat([trazabilidad, nueva_fila_trazabilidad], ignore_index=True)

                # Actualizar contadores y marcar como conciliados
                conteo_libro += 1
                conteo_banco += 1
                df_libro.at[idx_libro, 'Conciliado'] = True
                df_banco.at[idx_banco, 'Conciliado'] = True
                break

    return resultados, df_libro, df_banco, trazabilidad, conteo_libro, conteo_banco


def buscar_alertas(df_libro, df_banco, alerta):
    for idx_libro, partida_libro in df_libro[~df_libro['Conciliado']].iterrows():
        partidas_banco = df_banco[
            (abs(df_banco['Monto_Ajustado'] + partida_libro['Monto_Ajustado']) <= 0.05 ) &
            (df_banco['Fecha_Efectiva'].between(partida_libro['Fecha_Contable'] - timedelta(days=3), partida_libro['Fecha_Contable'] + timedelta(days=3))) &
            (~df_banco['Conciliado'])
        ]

        for idx_banco, partida_banco in partidas_banco.iterrows():
            aux = referencias_similares(partida_libro['Referencia_2'], partida_banco['Referencia_2'])
            if aux == 0:

                nueva_fila2 = pd.DataFrame({
                    'Partida_Libro': [partida_libro['Numero de Transacción']],
                    'Partida_Banco': [','.join(partidas_banco['Referencia'].astype(str))],
                    'Tipo_Conciliacion': ['conciliar_por_criterios_similares']
                })
                alerta = pd.concat([alerta, nueva_fila2], ignore_index=True)
                break

    return df_libro, df_banco, alerta


def verificar_integridad(df_libro, df_banco, resultados):
    total_libro = len(df_libro)
    total_banco = len(df_banco)
    conciliados_libro = df_libro['Conciliado'].sum()
    conciliados_banco = df_banco['Conciliado'].sum()
    no_conciliados_libro = total_libro - conciliados_libro
    no_conciliados_banco = total_banco - conciliados_banco

    print(f"Total registros en el libro: {total_libro}")
    print(f"Total registros en el banco: {total_banco}")
    print(f"Registros conciliados en el libro: {conciliados_libro}")
    print(f"Registros conciliados en el banco: {conciliados_banco}")
    print(f"Registros no conciliados en el libro: {no_conciliados_libro}")
    print(f"Registros no conciliados en el banco: {no_conciliados_banco}")

    if (conciliados_libro + no_conciliados_libro) != total_libro or (conciliados_banco + no_conciliados_banco) != total_banco:
        raise ValueError("Error en la conciliación: la suma de registros conciliados y no conciliados no coincide con el total de registros.")

def generar_excel_con_resultados(resultados, archivo_salida):
    resultados.to_excel(archivo_salida, index=False, engine='openpyxl')

    wb = load_workbook(archivo_salida)
    ws = wb.active

    fill_libro = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_banco = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[0].value == 'Libro':
            for cell in row:
                cell.fill = fill_libro
        elif row[0].value == 'Banco':
            for cell in row:
                cell.fill = fill_banco

    wb.save(archivo_salida)

def upload_file(filepath, mimetype, folder_id=None):
    """
    Upload a file to Google Drive.

    Args:
        filepath (str): The path to the file to upload.
        mimetype (str): The MIME type of the file.
        folder_id (str, optional): The ID of the folder to upload the file to. Defaults to None.

    Returns:
        str: The ID of the uploaded file.
    """
    filename = os.path.basename(filepath)
    file_metadata = {'name': filename}
    if folder_id:
        file_metadata['parents'] = [folder_id]
    media = MediaFileUpload(filepath, mimetype=mimetype)
    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    file_id = file.get('id')
    share_file_domain(service, file_id, "farmatodo.com")
    return file.get('id')

def share_file_domain(service, file_id, domain):
    """
    Share a file with all users in a specific domain.

    Args:
        service: The Google Drive service instance.
        file_id (str): The ID of the file to share.
        domain (str): The domain to share the file with (farmatodo.com).
    """
    permission = {
        'type': 'domain',
        'role': 'reader',
        'domain': domain
    }
    service.permissions().create(
        fileId=file_id,
        body=permission,
        sendNotificationEmail=False
    ).execute()

def update_appsheet_row(app_id, table_name, row_id, data):
    """
    Update a specific row in an AppSheet table.

    Args:
        app_id (str): The AppSheet app ID.
        table_name (str): The name of the table to update.
        row_id (str): The ID of the row to update.
        data (dict): The data to update the row with.
    """
    url = f"https://api.appsheet.com/api/v2/apps/{app_id}/tables/{table_name}/Action"
    
    headers = {
        "ApplicationAccessKey": "V2-OYuTL-e98Nu-wUK7e-Ha3or-TQXsi-CSCHV-Es6Cg-0r2aU",
        "Content-Type": "application/json"
    }
    
    payload = {
        "Action": "Edit",
        "Properties": {
            "Locale": "en-US",
            "TimeZone": "America/New_York"
        },
        "Rows": [{
            "Id": row_id,
            **data
        }]
    }
    
    response = requests.post(url, headers=headers, json=payload)
    print(f"Respuesta de AppSheet: {response.status_code}")
    print(f"Contenido de la respuesta: {response.text}")
    if response.status_code == 200:
        print(f"Row {row_id} {data} successfully updated in AppSheet")
    else:
        print(f"Failed to update row {row_id} in AppSheet. Status code: {response.status_code}")
        print(f"Response: {response.text}")

def add_appsheet_row(app_id, table_name, data):
    """
    Update a specific row in an AppSheet table.

    Args:
        app_id (str): The AppSheet app ID.
        table_name (str): The name of the table to update.
        row_id (str): The ID of the row to update.
        data (dict): The data to update the row with.
    """
    url = f"https://api.appsheet.com/api/v2/apps/{app_id}/tables/{table_name}/Action"
    
    headers = {
        "ApplicationAccessKey": "V2-OYuTL-e98Nu-wUK7e-Ha3or-TQXsi-CSCHV-Es6Cg-0r2aU",
        "Content-Type": "application/json"
    }
    
    payload = {
        "Action": "Add",
        "Properties": {
            "Locale": "en-US",
            "TimeZone": "America/New_York"
        },
        "Rows": [{
            **data
        }]
    }
    
    response = requests.post(url, headers=headers, json=payload)
    print(f"Respuesta de AppSheet: {response.status_code}")
    print(f"Contenido de la respuesta: {response.text}")
    if response.status_code == 200:
        print(f"{data} successfully updated in AppSheet")
    else:
        print(f"Failed to update in AppSheet. Status code: {response.status_code}")
        print(f"Response: {response.text}")


def to_price_format(value):
    """Convierte un valor a formato de precio con dos decimales."""
    return "{:.2f}".format(float(value))

def add_value_to_sheet(spreadsheet_id, sheet_name, data):
    body = {
        "values": [data]  # Asegúrate de que esto sea una lista de listas
    }
    # Cambiado a append para agregar filas
    sheets_service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range=sheet_name,  # Ajusta si tu hoja tiene un nombre diferente
        valueInputOption='USER_ENTERED',  # O 'RAW' según tus necesidades
        body=body
    ).execute()  # Este método ya agrega filas


@functions_framework.http
def conciliar_pagos_bancamiga_fondo_comun(request):
    try:
        print("Función iniciada")

        # VARIABLES
        # Result folders in Drive
        trazabilidad_folder_id = "16e16ATb8XkFlfMY0AgEwD1EqG1_5G3X6"
        resultados_folder_id = "1yrbCYcr1ZjDVzqQfsbi807Z_zDaIGUJ8"
        libro_no_conciliados_folder_id = "1se7QHI75TehbxsbDXlOvGu4-5TF4n_Du"
        banco_no_conciliados_folder_id = "16s1iVgZfGnW6sieN5krLKTjQUES2xCn_"
        alerta_registros_folder_id = "1gtqX54SjrvJl1IgmwMiL3rPRm-An-IWH"
        
        print("Creando carpetas locales")
        create_local_folders()
        
        print("Obteniendo datos de entrada")
        archivo_libro, archivo_banco, month, year, book_id = get_input_data(request)
        year = year.replace(",", "")

        print(f"Datos de entrada: libro={archivo_libro}, banco={archivo_banco}, mes={month}, año={year}, id={book_id}")

        # Files paths
        file1 = f"./tmp/libro_bruto/Libro_{month}_{year}_{book_id}.xlsx"
        file2 = f"./tmp/banco_bruto/Banco_{month}_{year}_{book_id}.xlsx"

        print("Descargando archivos de AppSheet")
        download_file_from_appsheet(archivo_libro, file1)
        download_file_from_appsheet(archivo_banco, file2)

        print("Iniciando proceso de conciliación")
        resultados, conteo_libro_conciliado, conteo_banco_conciliado, df_libro, df_banco, no_conciliados_libro, no_conciliados_banco, trazabilidad, conteo_por_criterio, total_banco_bruto, alerta,  conteo_banca_amiga, conteo_banco_fondo_comun = conciliar_pagos(file1, file2)

        print(f"Conciliación completada. Libro: {conteo_libro_conciliado}, Banco: {conteo_banco_conciliado}")

        print("Guardando resultados en archivos Excel")

        resultados.to_excel(f"./tmp/resultados/TOTAL_{month}_{year}_{book_id}.xlsx", index=False)
        trazabilidad.drop_duplicates().to_excel(f"./tmp/trazabilidad/TOTAL_{month}_{year}_{book_id}.xlsx", index=False)
        no_conciliados_libro.to_excel(f"./tmp/libro_no_conciliados/TOTAL_{month}_{year}_{book_id}.xlsx", index=False)
        no_conciliados_banco.to_excel(f"./tmp/banco_no_conciliados/TOTAL_{month}_{year}_{book_id}.xlsx", index=False)
        alerta.to_excel(f"./tmp/alerta/Alerta_{month}_{year}_{book_id}.xlsx", index=False)

        print("Subiendo archivos a Google Drive")
        trazabilidad_id = upload_file(f"./tmp/trazabilidad/TOTAL_{month}_{year}_{book_id}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", trazabilidad_folder_id)
        resultados_id = upload_file(f"./tmp/resultados/TOTAL_{month}_{year}_{book_id}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resultados_folder_id)
        libro_no_conciliados_id = upload_file(f"./tmp/libro_no_conciliados/TOTAL_{month}_{year}_{book_id}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", libro_no_conciliados_folder_id)
        banco_no_conciliados_id = upload_file(f"./tmp/banco_no_conciliados/TOTAL_{month}_{year}_{book_id}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", banco_no_conciliados_folder_id)
        alerta_doc_id = upload_file(f"./tmp/alerta/Alerta_{month}_{year}_{book_id}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", alerta_registros_folder_id)

        print("Actualizando AppSheet")
        app_id = "c539adf5-6f9d-42c1-9fe3-5f97f34e1034"
        
        # Asegúrate de que total_billed esté definido o cámbialo por un valor apropiado
        total_monto = df_libro['Monto'].abs().sum()  # Ajusta esto según la estructura real de tu DataFrame
        monto_conciliado = df_libro[df_libro['Conciliado']]['Monto'].abs().sum()

        total_conciliados = len(trazabilidad)
        
        print({
            "Archivo Resultado": f"https://drive.google.com/file/d/{resultados_id}",
            "Archivo de Conciliado": f"https://drive.google.com/file/d/{trazabilidad_id}",
            "Archivo de No Conciliados (Libro)": f"https://drive.google.com/file/d/{libro_no_conciliados_id}",
            "Archivo de No Conciliados (Banco": f"https://drive.google.com/file/d/{banco_no_conciliados_id}",
            "Cantidad de Transacciones": int(conteo_libro_conciliado) + int(conteo_banco_conciliado),
            "Cantidad de conciliados": int(total_conciliados),
            "Monto Total": to_price_format(total_monto),
            "Monto Conciliado":to_price_format(monto_conciliado),
            "Cantidad de Transacciones Libro": int(conteo_libro_conciliado),
            "Cantidad De Transacciones Banco": int(conteo_banco_conciliado),
            "Estado": "ACTUALIZADO",
            "Criterio 1 Banco": int(conteo_por_criterio['Criterio 1']['banco']),
            "Criterio 2 Banco": int(conteo_por_criterio['Criterio 2']['banco']),
            "Criterio 3 Banco": int(conteo_por_criterio['Criterio 3']['banco']),
            "Criterio 4 Banco": int(conteo_por_criterio['Criterio 4']['banco']),
            "Criterio 5 Banco": int(conteo_por_criterio['Criterio 5']['banco']),
            "Criterio 1 Libro": int(conteo_por_criterio['Criterio 1']['libro']),
            "Criterio 2 Libro": int(conteo_por_criterio['Criterio 2']['libro']),
            "Criterio 3 Libro": int(conteo_por_criterio['Criterio 3']['libro']),
            "Criterio 4 Libro": int(conteo_por_criterio['Criterio 4']['libro']),
            "Criterio 5 Libro": int(conteo_por_criterio['Criterio 5']['libro']),
            "Cantidad Total Banco": int(len(df_banco)),
            "Cantidad Total Libro": int(len(df_libro)),
            "Alerta": f"https://drive.google.com/file/d/{alerta_doc_id}",
            # "Conteo Banca Amiga": int(conteo_banca_amiga), 
            # "Conteo Fondo Comun": int(conteo_banco_fondo_comun),
        })

        update_appsheet_row(app_id, "Conciliación", book_id, {
            "Archivo Resultado": f"https://drive.google.com/file/d/{resultados_id}",
            "Archivo de Conciliado": f"https://drive.google.com/file/d/{trazabilidad_id}",
            "Archivo de No Conciliados (Libro)": f"https://drive.google.com/file/d/{libro_no_conciliados_id}",
            "Archivo de No Conciliados (Banco": f"https://drive.google.com/file/d/{banco_no_conciliados_id}",
            "Cantidad de Transacciones": int(conteo_libro_conciliado) + int(conteo_banco_conciliado),
            "Cantidad de conciliados": int(total_conciliados),
            "Monto Total": to_price_format(total_monto),
            "Monto Conciliado":to_price_format(monto_conciliado),
            "Cantidad de Transacciones Libro": int(conteo_libro_conciliado),
            "Cantidad De Transacciones Banco": int(conteo_banco_conciliado),
            "Estado": "ACTUALIZADO",
            "Criterio 1 Banco": int(conteo_por_criterio['Criterio 1']['banco']),
            "Criterio 2 Banco": int(conteo_por_criterio['Criterio 2']['banco']),
            "Criterio 3 Banco": int(conteo_por_criterio['Criterio 3']['banco']),
            "Criterio 4 Banco": int(conteo_por_criterio['Criterio 4']['banco']),
            "Criterio 5 Banco": int(conteo_por_criterio['Criterio 5']['banco']),
            "Criterio 1 Libro": int(conteo_por_criterio['Criterio 1']['libro']),
            "Criterio 2 Libro": int(conteo_por_criterio['Criterio 2']['libro']),
            "Criterio 3 Libro": int(conteo_por_criterio['Criterio 3']['libro']),
            "Criterio 4 Libro": int(conteo_por_criterio['Criterio 4']['libro']),
            "Criterio 5 Libro": int(conteo_por_criterio['Criterio 5']['libro']),
            "Cantidad Total Banco": int(len(df_banco)),
            "Cantidad Total Libro": int(len(df_libro)),
            "Alerta": f"https://drive.google.com/file/d/{alerta_doc_id}",
            "Conteo Banca Amiga": int(conteo_banca_amiga), 
            "Conteo Fondo Comun": int(conteo_banco_fondo_comun),
        })

        add_value_to_sheet("1KwxQUNs4dCHqYwt5gpkdJq21n93eTQvAcu2Zn9D2gbA","Detalles!A1",[
            book_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
            "Criterio 1",
            int(conteo_por_criterio['Criterio 1']['libro']),
            int(conteo_por_criterio['Criterio 1']['libro'])*100/int(conteo_libro_conciliado),
            int(conteo_por_criterio['Criterio 1']['banco']),
            int(conteo_por_criterio['Criterio 1']['banco'])*100/int(conteo_banco_conciliado)
        ])

        add_value_to_sheet("1KwxQUNs4dCHqYwt5gpkdJq21n93eTQvAcu2Zn9D2gbA","Detalles!A1",[
            book_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
            "Criterio 2",
            int(conteo_por_criterio['Criterio 2']['libro']),
            int(conteo_por_criterio['Criterio 2']['libro'])*100/int(conteo_libro_conciliado),
            int(conteo_por_criterio['Criterio 2']['banco']),
            int(conteo_por_criterio['Criterio 2']['banco'])*100/int(conteo_banco_conciliado)
        ])

        add_value_to_sheet("1KwxQUNs4dCHqYwt5gpkdJq21n93eTQvAcu2Zn9D2gbA","Detalles!A1",[
            book_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
            "Criterio 3",
            int(conteo_por_criterio['Criterio 3']['libro']),
            int(conteo_por_criterio['Criterio 3']['libro'])*100/int(conteo_libro_conciliado),
            int(conteo_por_criterio['Criterio 3']['banco']),
            int(conteo_por_criterio['Criterio 3']['banco'])*100/int(conteo_banco_conciliado)
        ])

        add_value_to_sheet("1KwxQUNs4dCHqYwt5gpkdJq21n93eTQvAcu2Zn9D2gbA","Detalles!A1",[
            book_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
            "Criterio 4",
            int(conteo_por_criterio['Criterio 4']['libro']),
            int(conteo_por_criterio['Criterio 4']['libro'])*100/int(conteo_libro_conciliado),
            int(conteo_por_criterio['Criterio 4']['banco']),
            int(conteo_por_criterio['Criterio 4']['banco'])*100/int(conteo_banco_conciliado)
        ])

        add_value_to_sheet("1KwxQUNs4dCHqYwt5gpkdJq21n93eTQvAcu2Zn9D2gbA","Detalles!A1",[
            book_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
            "Criterio 5",
            int(conteo_por_criterio['Criterio 5']['libro']),
            int(conteo_por_criterio['Criterio 5']['libro'])*100/int(conteo_libro_conciliado),
            int(conteo_por_criterio['Criterio 5']['banco']),
            int(conteo_por_criterio['Criterio 5']['banco'])*100/int(conteo_banco_conciliado)
        ])

        print("Función finalizada con éxito")
        return "Done", 200
    
    except Exception as e:
        print(f"Error en la función: {str(e)}")
        # Actualizar la hoja con el estatus de error
        update_appsheet_row(app_id, "Conciliación", book_id, {
             "Estado": "ERROR",
        })
        import traceback
        print(traceback.format_exc())
        return f"Error: {str(e)}", 500
